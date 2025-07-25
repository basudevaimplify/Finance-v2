"""
FastAPI Microservice for Enhanced AI Document Processing
Integrates with Finance Application's existing Node.js/React architecture
"""

import os
import json
import logging
import asyncio
from typing import Dict, List, Optional, Any
from datetime import datetime
import pandas as pd
import PyPDF2
import pdfplumber
from io import BytesIO, StringIO
import openpyxl
from openpyxl import load_workbook
import xlrd
import re

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Depends, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field, ConfigDict
import openai
from openai import OpenAI
import asyncpg
from contextlib import asynccontextmanager
import uuid
import hashlib
import bcrypt
from pathlib import Path
import shutil

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Database configuration
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:1@localhost:5432/finance_app")
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Security
security = HTTPBearer()
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-here")

# Global database connection pool
db_pool = None

# Utility Functions
def _to_camel_case(s: str) -> str:
    """Convert snake_case string to camelCase."""
    parts = s.split('_')
    return parts[0] + ''.join(p.capitalize() for p in parts[1:])

def record_to_camel(record: Any) -> Dict[str, Any]:
    """Convert an asyncpg Record to a camelCase keyed dict."""
    return {_to_camel_case(k): v for k, v in dict(record).items()}

# Database Models
class User(BaseModel):
    id: str
    email: str
    first_name: str
    last_name: str
    tenant_id: Optional[str] = None

    @classmethod
    def from_db_row(cls, row):
        """Create User from database row, converting UUID to string"""
        return cls(
            id=row['id'],
            email=row['email'],
            first_name=row['first_name'],
            last_name=row['last_name'],
            tenant_id=str(row['tenant_id']) if row['tenant_id'] else None
        )

class Document(BaseModel):
    id: str
    file_name: str
    original_name: str
    mime_type: str
    file_size: int
    file_path: str
    document_type: str
    status: str
    uploaded_by: str
    tenant_id: str
    metadata: Dict[str, Any]
    extracted_data: Optional[Dict[str, Any]] = None
    created_at: datetime
    updated_at: datetime

    model_config = ConfigDict(alias_generator=_to_camel_case, populate_by_name=True)

class DocumentUploadResponse(BaseModel):
    document: Document
    message: str
    processing_time_ms: int

    model_config = ConfigDict(alias_generator=_to_camel_case, populate_by_name=True)

# Authentication Models
class LoginRequest(BaseModel):
    email: str
    password: str

class LoginResponse(BaseModel):
    token: str
    user: User

# Database Functions
async def init_db():
    """Initialize database connection pool"""
    global db_pool
    try:
        db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=10)
        logger.info("✅ Database connection pool created")
    except Exception as e:
        logger.error(f"❌ Failed to create database pool: {e}")
        raise

async def close_db():
    """Close database connection pool"""
    global db_pool
    if db_pool:
        await db_pool.close()
        logger.info("Database connection pool closed")

async def get_db():
    """Get database connection from pool"""
    if not db_pool:
        raise HTTPException(status_code=500, detail="Database not initialized")
    return db_pool

# Authentication Functions
def hash_password(password: str) -> str:
    """Hash password using bcrypt"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    """Verify password against hash"""
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def generate_token(user_id: str) -> str:
    """Generate simple token (in production, use JWT)"""
    return hashlib.sha256(f"{user_id}{SECRET_KEY}".encode()).hexdigest()



async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security), db = Depends(get_db)) -> User:
    """Get current authenticated user"""
    token = credentials.credentials

    async with db.acquire() as conn:
        user_row = await conn.fetchrow(
            "SELECT id, email, first_name, last_name, tenant_id FROM users WHERE token = $1",
            token
        )

        if not user_row:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid authentication credentials"
            )

        return User.from_db_row(user_row)

# Database lifecycle management
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    await init_db()
    yield
    # Shutdown
    await close_db()

# Initialize FastAPI app
app = FastAPI(
    title="Finance AI Document Processor",
    description="Enhanced AI document processing and backend services",
    version="2.0.0",
    lifespan=lifespan
)

# Configure CORS for direct React frontend communication
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:8000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files for React frontend and uploaded files
app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# Initialize OpenAI client
openai_client = None
if os.getenv("OPENAI_API_KEY"):
    openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    logger.info("OpenAI client initialized successfully")
else:
    logger.warning("OpenAI API key not found. AI features will be limited.")

# Pydantic models for API responses
class ClassificationResult(BaseModel):
    document_type: str = Field(..., description="Classified document type")
    confidence: float = Field(..., description="Classification confidence (0-1)")
    reasoning: str = Field(..., description="Explanation for classification")
    key_indicators: List[str] = Field(default=[], description="Key indicators found")
    content_summary: str = Field(..., description="Brief content summary")
    potential_misclassification: bool = Field(default=False, description="Flag for uncertain classification")

class ExtractedRecord(BaseModel):
    row_index: int
    data: Dict[str, Any]
    confidence: float = Field(default=1.0)

class ExtractionResult(BaseModel):
    records: List[ExtractedRecord]
    total_records: int
    extraction_confidence: float
    schema_detected: Dict[str, str]
    processing_notes: List[str] = Field(default=[])

class DocumentProcessingResponse(BaseModel):
    classification: ClassificationResult
    extraction: ExtractionResult
    processing_time_ms: int
    ai_enhanced: bool = Field(default=False)
    raw_text_content: str = Field(default="")

# Document type classification prompts
CLASSIFICATION_PROMPTS = {
    "system": """You are an expert financial document classifier. Analyze the provided document content and classify it into one of these types:
    
    1. sales_register - Contains sales/invoice data with customer information, amounts, GST details
    2. purchase_register - Contains purchase/vendor data with supplier information, amounts, tax details
    3. bank_statement - Contains bank transaction data with dates, descriptions, debit/credit amounts
    4. other - Any other type of document
    
    Provide a confidence score (0-1) and detailed reasoning for your classification.""",
    
    "user_template": """Analyze this document content and classify it:

Content Preview:
{content_preview}

File Name: {filename}
File Type: {file_type}

Respond with a JSON object containing:
- document_type: one of [sales_register, purchase_register, bank_statement, other]
- confidence: float between 0 and 1
- reasoning: detailed explanation
- key_indicators: list of specific indicators found
- content_summary: brief summary of content
- potential_misclassification: boolean if confidence < 0.8"""
}

# Data extraction prompts for different document types
EXTRACTION_PROMPTS = {
    "sales_register": {
        "system": "Extract structured sales data from the document. Focus on invoice numbers, customer details, amounts, dates, and GST information.",
        "schema": {
            "invoice_no": "Invoice/Bill number",
            "date": "Invoice date",
            "customer_name": "Customer/Client name",
            "customer_gstin": "Customer GSTIN",
            "item_description": "Product/Service description",
            "hsn_code": "HSN/SAC code",
            "quantity": "Quantity",
            "rate": "Rate per unit",
            "taxable_value": "Taxable amount",
            "cgst_rate": "CGST rate",
            "cgst_amount": "CGST amount",
            "sgst_rate": "SGST rate", 
            "sgst_amount": "SGST amount",
            "igst_rate": "IGST rate",
            "igst_amount": "IGST amount",
            "total_amount": "Total invoice amount"
        }
    },
    "purchase_register": {
        "system": "Extract structured purchase data from the document. Focus on vendor details, purchase amounts, tax information, and TDS details.",
        "schema": {
            "purchase_order_no": "Purchase order number",
            "date": "Purchase date",
            "vendor_name": "Vendor/Supplier name",
            "vendor_gstin": "Vendor GSTIN",
            "bill_no": "Bill/Invoice number",
            "item_description": "Product/Service description",
            "hsn_code": "HSN/SAC code",
            "quantity": "Quantity",
            "rate": "Rate per unit",
            "taxable_value": "Taxable amount",
            "cgst_rate": "CGST rate",
            "cgst_amount": "CGST amount",
            "sgst_rate": "SGST rate",
            "sgst_amount": "SGST amount", 
            "igst_rate": "IGST rate",
            "igst_amount": "IGST amount",
            "total_amount": "Total amount",
            "tds_rate": "TDS rate",
            "tds_amount": "TDS amount",
            "net_payable": "Net payable amount"
        }
    },
    "bank_statement": {
        "system": "Extract structured bank transaction data from the document. Focus on transaction dates, descriptions, amounts, and balances.",
        "schema": {
            "date": "Transaction date",
            "transaction_id": "Transaction ID/Reference",
            "description": "Transaction description",
            "reference_no": "Reference number",
            "debit_amount": "Debit amount (money going out)",
            "credit_amount": "Credit amount (money coming in)",
            "balance": "Account balance after transaction"
        }
    }
}

# File format detection and validation
SUPPORTED_FILE_TYPES = {
    'application/pdf': ['.pdf'],
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': ['.xlsx'],
    'application/vnd.ms-excel': ['.xls'],
    'text/csv': ['.csv'],
    'application/octet-stream': ['.csv', '.xlsx', '.xls', '.pdf']  # Fallback for unclear MIME types
}

def detect_file_format(file: UploadFile) -> str:
    """Detect file format based on extension and MIME type"""
    filename = file.filename.lower() if file.filename else ""

    if filename.endswith('.pdf'):
        return 'pdf'
    elif filename.endswith(('.xlsx', '.xls')):
        return 'excel'
    elif filename.endswith('.csv'):
        return 'csv'
    else:
        # Fallback to MIME type detection
        if file.content_type == 'application/pdf':
            return 'pdf'
        elif file.content_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
            return 'excel'
        elif file.content_type == 'text/csv':
            return 'csv'
        else:
            return 'unknown'

async def extract_text_from_pdf(content: bytes) -> str:
    """Extract text from PDF using pdfplumber for better accuracy"""
    try:
        text_content = []

        # Try pdfplumber first (better for structured documents)
        try:
            with pdfplumber.open(BytesIO(content)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_content.append(page_text)
        except Exception as e:
            logger.warning(f"pdfplumber failed, trying PyPDF2: {e}")

            # Fallback to PyPDF2
            pdf_reader = PyPDF2.PdfReader(BytesIO(content))
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_content.append(page_text)

        return "\n".join(text_content)
    except Exception as e:
        logger.error(f"PDF text extraction failed: {e}")
        return ""

async def extract_data_from_excel(content: bytes, filename: str) -> Dict[str, Any]:
    """Extract structured data from Excel files"""
    try:
        # Try openpyxl for .xlsx files
        if filename.endswith('.xlsx'):
            workbook = load_workbook(BytesIO(content), read_only=True)
            sheet = workbook.active

            # Extract headers and data
            headers = []
            data_rows = []

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
                else:
                    if any(cell is not None for cell in row):
                        data_rows.append([str(cell) if cell is not None else "" for cell in row])

            return {
                'headers': headers,
                'data': data_rows,
                'format': 'excel',
                'sheet_count': len(workbook.sheetnames)
            }

        # Try pandas for .xls files
        else:
            df = pd.read_excel(BytesIO(content))
            return {
                'headers': df.columns.tolist(),
                'data': df.values.tolist(),
                'format': 'excel',
                'sheet_count': 1
            }

    except Exception as e:
        logger.error(f"Excel data extraction failed: {e}")
        return {'headers': [], 'data': [], 'format': 'excel', 'error': str(e)}

async def extract_text_from_file(file: UploadFile) -> str:
    """Enhanced text extraction with multi-format support"""
    try:
        content = await file.read()
        file_format = detect_file_format(file)

        if file_format == 'pdf':
            # Use enhanced PDF extraction
            return await extract_text_from_pdf(content)

        elif file_format == 'excel':
            # Extract Excel data and convert to text representation
            excel_data = await extract_data_from_excel(content, file.filename or "")
            if excel_data.get('error'):
                raise HTTPException(status_code=400, detail=f"Excel processing failed: {excel_data['error']}")

            # Convert structured data to text for classification
            text_lines = []
            if excel_data['headers']:
                text_lines.append(','.join(excel_data['headers']))

            for row in excel_data['data'][:50]:  # Limit to first 50 rows for classification
                text_lines.append(','.join(str(cell) for cell in row))

            return '\n'.join(text_lines)

        elif file_format == 'csv':
            # Read CSV content
            try:
                text_content = content.decode('utf-8')
            except UnicodeDecodeError:
                text_content = content.decode('latin-1')  # Fallback encoding
            return text_content

        else:
            # Try to decode as text for unknown formats
            try:
                return content.decode('utf-8')
            except UnicodeDecodeError:
                return content.decode('latin-1')

    except Exception as e:
        logger.error(f"Error extracting text from file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to extract text from file: {str(e)}")

async def classify_document_with_ai(content: str, filename: str, file_type: str) -> ClassificationResult:
    """Classify document using OpenAI GPT-4"""
    if not openai_client:
        # Fallback classification based on filename and content patterns
        return fallback_classification(content, filename, file_type)
    
    try:
        # Prepare content preview (first 2000 characters)
        content_preview = content[:2000] if len(content) > 2000 else content
        
        user_prompt = CLASSIFICATION_PROMPTS["user_template"].format(
            content_preview=content_preview,
            filename=filename,
            file_type=file_type
        )
        
        response = openai_client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": CLASSIFICATION_PROMPTS["system"]},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1,
            max_tokens=500
        )
        
        # Parse AI response
        ai_response = response.choices[0].message.content
        
        try:
            # Try to parse as JSON
            result_data = json.loads(ai_response)
            return ClassificationResult(**result_data)
        except json.JSONDecodeError:
            # If not JSON, parse manually
            logger.warning("AI response not in JSON format, using fallback parsing")
            return fallback_classification(content, filename, file_type)
            
    except Exception as e:
        logger.error(f"Error in AI classification: {e}")
        return fallback_classification(content, filename, file_type)

def fallback_classification(content: str, filename: str, file_type: str) -> ClassificationResult:
    """Fallback classification without AI"""
    content_lower = content.lower()
    filename_lower = filename.lower()
    
    # Check for bank statement indicators
    bank_indicators = ["transaction", "debit", "credit", "balance", "bank", "statement", "account"]
    bank_score = sum(1 for indicator in bank_indicators if indicator in content_lower or indicator in filename_lower)
    
    # Check for sales register indicators  
    sales_indicators = ["invoice", "customer", "sales", "gst", "cgst", "sgst", "bill"]
    sales_score = sum(1 for indicator in sales_indicators if indicator in content_lower or indicator in filename_lower)
    
    # Check for purchase register indicators
    purchase_indicators = ["purchase", "vendor", "supplier", "po", "tds", "payable"]
    purchase_score = sum(1 for indicator in purchase_indicators if indicator in content_lower or indicator in filename_lower)
    
    # Determine classification
    max_score = max(bank_score, sales_score, purchase_score)
    
    if max_score == 0:
        return ClassificationResult(
            document_type="other",
            confidence=0.2,
            reasoning="No clear financial document indicators found",
            key_indicators=["analysis_unclear"],
            content_summary="Document type uncertain"
        )
    
    if bank_score == max_score:
        doc_type = "bank_statement"
        confidence = min(0.9, 0.5 + (bank_score * 0.1))
        indicators = [ind for ind in bank_indicators if ind in content_lower or ind in filename_lower]
    elif sales_score == max_score:
        doc_type = "sales_register"
        confidence = min(0.9, 0.5 + (sales_score * 0.1))
        indicators = [ind for ind in sales_indicators if ind in content_lower or ind in filename_lower]
    else:
        doc_type = "purchase_register"
        confidence = min(0.9, 0.5 + (purchase_score * 0.1))
        indicators = [ind for ind in purchase_indicators if ind in content_lower or ind in filename_lower]
    
    return ClassificationResult(
        document_type=doc_type,
        confidence=confidence,
        reasoning=f"Pattern-based classification found {max_score} indicators for {doc_type}",
        key_indicators=indicators,
        content_summary=f"Document appears to be a {doc_type.replace('_', ' ')}",
        potential_misclassification=confidence < 0.8
    )

async def extract_data_with_ai(content: str, doc_type: str) -> ExtractionResult:
    """Extract structured data using AI or fallback methods"""
    logger.info(f"🔍 Starting data extraction for document type: {doc_type}")

    if doc_type not in EXTRACTION_PROMPTS:
        logger.warning(f"❌ Unsupported document type: {doc_type}")
        return ExtractionResult(
            records=[],
            total_records=0,
            extraction_confidence=0.0,
            schema_detected={},
            processing_notes=["Unsupported document type for extraction"]
        )

    logger.info(f"📊 Content length: {len(content)} characters")

    # Try CSV parsing first for structured data
    try:
        logger.info("📋 Attempting CSV parsing...")
        # Parse as CSV
        df = pd.read_csv(StringIO(content))
        logger.info(f"✅ CSV parsed successfully: {len(df)} rows, {len(df.columns)} columns")
        logger.info(f"📝 Columns detected: {list(df.columns)}")

        if len(df) > 0:
            records = []
            schema_info = EXTRACTION_PROMPTS[doc_type]["schema"]
            logger.info(f"🔧 Processing {len(df)} rows...")

            for idx, row in df.iterrows():
                record_data = {}
                for col in df.columns:
                    record_data[col] = row[col] if pd.notna(row[col]) else None

                records.append(ExtractedRecord(
                    row_index=idx,
                    data=record_data,
                    confidence=0.95
                ))

            logger.info(f"✅ Successfully extracted {len(records)} records from CSV")
            return ExtractionResult(
                records=records,
                total_records=len(records),
                extraction_confidence=0.95,
                schema_detected={col: str(df[col].dtype) for col in df.columns},
                processing_notes=["Successfully parsed as CSV format"]
            )

    except Exception as e:
        logger.info(f"⚠️ CSV parsing failed, trying alternative extraction: {e}")

    # Fallback to basic text parsing
    logger.info("🔧 Falling back to basic text parsing...")
    lines = content.strip().split('\n')
    records = []
    logger.info(f"📄 Processing {len(lines)} lines...")

    for idx, line in enumerate(lines[1:], 1):  # Skip header
        if line.strip():
            # Basic comma-separated parsing
            values = [v.strip() for v in line.split(',')]
            if len(values) > 1:
                record_data = {f"field_{i}": val for i, val in enumerate(values)}
                records.append(ExtractedRecord(
                    row_index=idx,
                    data=record_data,
                    confidence=0.7
                ))

    logger.info(f"✅ Basic parsing completed: {len(records)} records extracted")
    
    return ExtractionResult(
        records=records,
        total_records=len(records),
        extraction_confidence=0.7,
        schema_detected={"format": "text_parsed"},
        processing_notes=["Used fallback text parsing"]
    )

# async def process_bank_statement_pdf(content: bytes, filename: str) -> ExtractionResult:
#     """Process bank statement PDF with enhanced transaction extraction"""
#     try:
#         text = await extract_text_from_pdf(content)
#         logger.info(f"📄 PDF text extracted: {len(text)} characters")

#         # Enhanced bank statement patterns for different formats
#         transaction_patterns = [
#             # Standard format: Date Description Reference Debit Credit Balance
#             r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(.+?)\s+([A-Z0-9]+)\s+(\d+[,.]?\d*\.?\d*)\s+(\d+[,.]?\d*\.?\d*)',

#             # Format with optional debit/credit: Date Description Amount Balance
#             r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(.+?)\s+(\d+[,.]?\d*\.?\d*)\s+(\d+[,.]?\d*\.?\d*)',

#             # Format with reference: Date Description Reference Amount
#             r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(.+?)\s+([A-Z0-9]+)\s+(\d+[,.]?\d*\.?\d*)',

#             # Simple format: Date Description Amount
#             r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(.+?)\s+(\d+[,.]?\d*\.?\d*)',

#             # Tab-separated format
#             r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\t+(.+?)\t+(.+?)\t+(\d+[,.]?\d*\.?\d*)',
#         ]

#         records = []
#         processed_lines = []

#         # Process each line
#         for line_num, line in enumerate(text.split('\n')):
#             line = line.strip()
#             if not line or len(line) < 10:  # Skip empty or very short lines
#                 continue

#             # Skip header lines
#             if any(header in line.lower() for header in ['date', 'description', 'reference', 'debit', 'credit', 'balance', 'transaction', 'account']):
#                 continue

#             # Try each pattern
#             for pattern_idx, pattern in enumerate(transaction_patterns):
#                 match = re.search(pattern, line)
#                 if match:
#                     groups = match.groups()
#                     logger.info(f"🔍 Pattern {pattern_idx} matched line {line_num}: {groups}")

#                     # Extract data based on number of groups and analyze the line structure
#                     if len(groups) >= 3:
#                         # Analyze the line to determine the structure
#                         # Look for keywords to determine if it's debit or credit
#                         line_lower = line.lower()
#                         is_credit = any(keyword in line_lower for keyword in ['payment received', 'deposit', 'revenue', 'income', 'credit', 'retainer'])
#                         is_debit = any(keyword in line_lower for keyword in ['payment', 'expense', 'charges', 'purchase', 'bills', 'premium']) and not is_credit

#                         if len(groups) >= 5:  # Date, Description, Reference, Amount1, Amount2 (could be Debit/Credit or Amount/Balance)
#                             # Check if we have separate debit/credit columns or amount/balance
#                             amount1 = groups[3].replace(',', '')
#                             amount2 = groups[4].replace(',', '')

#                             # If both amounts exist, likely Amount and Balance format
#                             if amount1 and amount2:
#                                 record_data = {
#                                     'Date': groups[0],
#                                     'Description': groups[1].strip(),
#                                     'Reference': groups[2],
#                                     'Debit Amount': amount1 if is_debit else None,
#                                     'Credit Amount': amount1 if is_credit else None,
#                                     'Balance': amount2
#                                 }
#                             else:
#                                 # One is debit, one is credit
#                                 record_data = {
#                                     'Date': groups[0],
#                                     'Description': groups[1].strip(),
#                                     'Reference': groups[2],
#                                     'Debit Amount': amount1 if amount1 and amount1 != '0' else None,
#                                     'Credit Amount': amount2 if amount2 and amount2 != '0' else None,
#                                     'Balance': None
#                                 }
#                         elif len(groups) == 4:  # Date, Description, Reference, Amount OR Date, Description, Amount, Balance
#                             # Check if third group looks like a reference (letters/numbers) or amount (only numbers)
#                             if re.match(r'^[A-Z0-9]+$', groups[2]):  # Has reference
#                                 amount = groups[3].replace(',', '')
#                                 record_data = {
#                                     'Date': groups[0],
#                                     'Description': groups[1].strip(),
#                                     'Reference': groups[2],
#                                     'Debit Amount': amount if is_debit else None,
#                                     'Credit Amount': amount if is_credit else None,
#                                     'Balance': None
#                                 }
#                             else:  # Amount, Balance format
#                                 amount = groups[2].replace(',', '')
#                                 balance = groups[3].replace(',', '')
#                                 record_data = {
#                                     'Date': groups[0],
#                                     'Description': groups[1].strip(),
#                                     'Reference': None,
#                                     'Debit Amount': amount if is_debit else None,
#                                     'Credit Amount': amount if is_credit else None,
#                                     'Balance': balance
#                                 }
#                         else:  # Date, Description, Amount
#                             amount = groups[2].replace(',', '')
#                             record_data = {
#                                 'Date': groups[0],
#                                 'Description': groups[1].strip(),
#                                 'Reference': None,
#                                 'Debit Amount': amount if is_debit else None,
#                                 'Credit Amount': amount if is_credit else None,
#                                 'Balance': None
#                             }

#                         # Add confidence and source
#                         record_data['confidence'] = 0.85
#                         record_data['source'] = 'pdf_extraction'
#                         record_data['id'] = f"pdf-record-{len(records)}"

#                         records.append(ExtractedRecord(
#                             row_index=len(records),
#                             data=record_data,
#                             confidence=0.85
#                         ))

#                         processed_lines.append(f"Line {line_num}: {line}")
#                         break  # Stop trying other patterns for this line

#         logger.info(f"📊 PDF processing completed: {len(records)} records extracted from {len(processed_lines)} lines")

#         return ExtractionResult(
#             records=records,
#             total_records=len(records),
#             extraction_confidence=0.85 if records else 0.3,
#             schema_detected={'format': 'bank_statement_pdf', 'patterns_used': str(len(transaction_patterns))},
#             processing_notes=[f"Extracted {len(records)} transactions from PDF using enhanced patterns"]
#         )

#     except Exception as e:
#         logger.error(f"Bank statement PDF processing failed: {e}")
#         return ExtractionResult(
#             records=[],
#             total_records=0,
#             extraction_confidence=0.0,
#             schema_detected={},
#             processing_notes=[f"PDF processing failed: {str(e)}"]
#         )
# ============================================================================
# UNIVERSAL AI-POWERED PDF FINANCIAL DOCUMENT PROCESSOR
# ============================================================================

from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple, Union
from enum import Enum
import re
from datetime import datetime

class FieldType(Enum):
    """Semantic field types for financial documents"""
    DATE = "date"
    DESCRIPTION = "description"
    REFERENCE = "reference"
    AMOUNT = "amount"
    DEBIT = "debit"
    CREDIT = "credit"
    BALANCE = "balance"
    ACCOUNT = "account"
    INVOICE_NUMBER = "invoice_number"
    CUSTOMER = "customer"
    VENDOR = "vendor"
    TAX = "tax"
    UNKNOWN = "unknown"

@dataclass
class TableRegion:
    """Represents a detected table region in the document"""
    start_line: int
    end_line: int
    lines: List[str]
    confidence: float
    table_type: str  # 'structured', 'semi_structured', 'key_value'

@dataclass
class ColumnStructure:
    """Represents detected column structure"""
    positions: List[int]
    headers: List[str]
    alignment: str  # 'fixed_width', 'tab_separated', 'space_separated'
    confidence: float

@dataclass
class FieldMapping:
    """Maps detected fields to standard schema"""
    original_name: str
    field_type: FieldType
    standard_name: str
    confidence: float
    sample_values: List[str]

class PDFIntelligenceEngine:
    """Core intelligence engine for PDF document analysis"""

    def __init__(self):
        self.date_patterns = [
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
            r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',
            r'\d{1,2}\s+[A-Za-z]{3}\s+\d{4}',
            r'[A-Za-z]{3}\s+\d{1,2},?\s+\d{4}',
        ]

        self.amount_patterns = [
            r'₹\s*[\d,]+\.?\d*',
            r'\$\s*[\d,]+\.?\d*',
            r'€\s*[\d,]+\.?\d*',
            r'Rs\.?\s*[\d,]+\.?\d*',
            r'[\d,]+\.?\d*',
            r'\([\d,]+\.?\d*\)',  # Negative amounts in parentheses
        ]

        self.reference_patterns = [
            r'[A-Z]{2,4}\d{3,}',
            r'INV[/-]?\d+',
            r'CHQ\d+',
            r'TXN\d+',
            r'AUTO\d+',
            r'DEP\d+',
            r'[A-Z]+\d+',
        ]

    def classify_document_type(self, text: str) -> Tuple[str, float]:
        """Classify document type based on content analysis"""
        text_lower = text.lower()

        # Bank statement indicators
        bank_indicators = ['bank', 'statement', 'account', 'balance', 'transaction', 'debit', 'credit']
        bank_score = sum(1 for indicator in bank_indicators if indicator in text_lower)

        # Invoice indicators
        invoice_indicators = ['invoice', 'bill', 'amount due', 'payment terms', 'invoice number']
        invoice_score = sum(1 for indicator in invoice_indicators if indicator in text_lower)

        # Sales register indicators
        sales_indicators = ['sales', 'customer', 'invoice', 'gst', 'tax', 'total amount']
        sales_score = sum(1 for indicator in sales_indicators if indicator in text_lower)

        # Purchase register indicators
        purchase_indicators = ['purchase', 'vendor', 'supplier', 'bill', 'gst', 'tax']
        purchase_score = sum(1 for indicator in purchase_indicators if indicator in text_lower)

        scores = {
            'bank_statement': bank_score,
            'invoice': invoice_score,
            'sales_register': sales_score,
            'purchase_register': purchase_score
        }

        doc_type = max(scores, key=scores.get)
        confidence = min(scores[doc_type] / 10.0, 1.0)  # Normalize to 0-1

        return doc_type, confidence

    def detect_field_type(self, header: str, sample_values: List[str]) -> FieldType:
        """Detect field type based on header name and sample values"""
        header_lower = header.lower().strip()

        # Date field detection
        if any(keyword in header_lower for keyword in ['date', 'dt', 'time']):
            return FieldType.DATE

        # Amount field detection
        if any(keyword in header_lower for keyword in ['amount', 'amt', 'value', 'total']):
            return FieldType.AMOUNT

        # Debit field detection
        if any(keyword in header_lower for keyword in ['debit', 'dr', 'withdrawal', 'expense']):
            return FieldType.DEBIT

        # Credit field detection
        if any(keyword in header_lower for keyword in ['credit', 'cr', 'deposit', 'income']):
            return FieldType.CREDIT

        # Balance field detection
        if any(keyword in header_lower for keyword in ['balance', 'bal', 'closing']):
            return FieldType.BALANCE

        # Reference field detection
        if any(keyword in header_lower for keyword in ['reference', 'ref', 'cheque', 'chq', 'txn']):
            return FieldType.REFERENCE

        # Description field detection
        if any(keyword in header_lower for keyword in ['description', 'desc', 'particulars', 'narration']):
            return FieldType.DESCRIPTION

        # Customer/Vendor detection
        if any(keyword in header_lower for keyword in ['customer', 'client', 'party']):
            return FieldType.CUSTOMER
        if any(keyword in header_lower for keyword in ['vendor', 'supplier']):
            return FieldType.VENDOR

        # Invoice number detection
        if any(keyword in header_lower for keyword in ['invoice', 'bill', 'voucher']):
            return FieldType.INVOICE_NUMBER

        # Analyze sample values for pattern detection
        if sample_values:
            # Check if values look like dates
            date_count = sum(1 for val in sample_values[:5] if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', str(val)))
            if date_count > len(sample_values) * 0.6:
                return FieldType.DATE

            # Check if values look like amounts
            amount_count = sum(1 for val in sample_values[:5] if re.search(r'[\d,]+\.?\d*', str(val)))
            if amount_count > len(sample_values) * 0.6:
                return FieldType.AMOUNT

        return FieldType.UNKNOWN

async def process_pdf_document(content: bytes, filename: str) -> ExtractionResult:
    """Universal PDF document processor - handles ANY financial document format"""
    try:
        # Initialize the intelligence engine
        engine = PDFIntelligenceEngine()

        # Extract text from PDF
        text = await extract_text_from_pdf(content)
        logger.info(f"📄 PDF text extracted: {len(text)} characters")

        # Stage 1: Document type classification
        doc_type, type_confidence = engine.classify_document_type(text)
        logger.info(f"🔍 Document classified as: {doc_type} (confidence: {type_confidence:.2f})")

        # Stage 2: Table structure detection
        detector = TableStructureDetector()
        table_regions = detector.detect_table_boundaries(text)
        logger.info(f"📊 Detected {len(table_regions)} table regions")

        if not table_regions:
            logger.warning("⚠️ No table regions detected, falling back to line-by-line processing")
            return await _fallback_line_processing(text, engine)

        # Process the best table region
        best_table = max(table_regions, key=lambda t: t.confidence)
        logger.info(f"🎯 Processing best table region (confidence: {best_table.confidence:.2f}, type: {best_table.table_type})")

        # Stage 3: Dynamic header and column discovery
        mapper = SemanticFieldMapper(engine)
        field_mappings = await mapper.discover_fields(best_table.lines)
        logger.info(f"🗂️ Discovered {len(field_mappings)} field mappings")

        # Stage 4: Data extraction with the discovered structure
        orchestrator = DataExtractionOrchestrator(engine, mapper)
        extraction_result = await orchestrator.extract_data(best_table, field_mappings)

        logger.info(f"✅ Universal PDF processing completed: {extraction_result.total_records} records extracted")
        return extraction_result

    except Exception as e:
        logger.error(f"Universal PDF processing failed: {e}")
        return ExtractionResult(
            records=[],
            total_records=0,
            extraction_confidence=0.0,
            schema_detected={},
            processing_notes=[f"Universal PDF processing failed: {str(e)}"]
        )

class TableStructureDetector:
    """Detects and analyzes table structures in PDF text"""

    def __init__(self):
        self.min_table_lines = 3
        self.min_columns = 2

    def detect_table_boundaries(self, text: str) -> List[TableRegion]:
        """Detect table regions in the document"""
        lines = text.split('\n')
        table_regions = []
        current_table_start = None
        current_table_lines = []

        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                if current_table_start is not None and len(current_table_lines) >= self.min_table_lines:
                    # End of table region
                    confidence = self._calculate_table_confidence(current_table_lines)
                    table_type = self._determine_table_type(current_table_lines)

                    table_regions.append(TableRegion(
                        start_line=current_table_start,
                        end_line=i-1,
                        lines=current_table_lines.copy(),
                        confidence=confidence,
                        table_type=table_type
                    ))

                current_table_start = None
                current_table_lines = []
                continue

            # Check if line looks like table data
            if self._is_table_line(line):
                if current_table_start is None:
                    current_table_start = i
                current_table_lines.append(line)
            else:
                # Check if this might be a header line
                if current_table_start is None and self._is_header_line(line):
                    current_table_start = i
                    current_table_lines.append(line)
                elif current_table_start is not None:
                    # End current table if we have enough lines
                    if len(current_table_lines) >= self.min_table_lines:
                        confidence = self._calculate_table_confidence(current_table_lines)
                        table_type = self._determine_table_type(current_table_lines)

                        table_regions.append(TableRegion(
                            start_line=current_table_start,
                            end_line=i-1,
                            lines=current_table_lines.copy(),
                            confidence=confidence,
                            table_type=table_type
                        ))

                    current_table_start = None
                    current_table_lines = []

        # Handle table at end of document
        if current_table_start is not None and len(current_table_lines) >= self.min_table_lines:
            confidence = self._calculate_table_confidence(current_table_lines)
            table_type = self._determine_table_type(current_table_lines)

            table_regions.append(TableRegion(
                start_line=current_table_start,
                end_line=len(lines)-1,
                lines=current_table_lines,
                confidence=confidence,
                table_type=table_type
            ))

        return table_regions

    def _is_table_line(self, line: str) -> bool:
        """Check if a line appears to be table data"""
        # Look for patterns indicating tabular data
        has_date = bool(re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', line))
        has_amount = bool(re.search(r'[\d,]+\.?\d*', line))
        has_multiple_fields = len(line.split()) >= 3
        has_tabs = '\t' in line

        return (has_date and has_amount) or has_tabs or (has_multiple_fields and has_amount)

    def _is_header_line(self, line: str) -> bool:
        """Check if a line appears to be a table header"""
        header_keywords = ['date', 'description', 'amount', 'debit', 'credit', 'balance',
                          'reference', 'invoice', 'customer', 'vendor', 'tax']
        line_lower = line.lower()
        return sum(1 for keyword in header_keywords if keyword in line_lower) >= 2

    def _calculate_table_confidence(self, lines: List[str]) -> float:
        """Calculate confidence that these lines form a valid table"""
        if not lines:
            return 0.0

        # Factors that increase confidence
        date_lines = sum(1 for line in lines if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', line))
        amount_lines = sum(1 for line in lines if re.search(r'[\d,]+\.?\d*', line))
        consistent_columns = self._check_column_consistency(lines)

        date_ratio = date_lines / len(lines)
        amount_ratio = amount_lines / len(lines)

        confidence = (date_ratio * 0.4 + amount_ratio * 0.4 + consistent_columns * 0.2)
        return min(confidence, 1.0)

    def _determine_table_type(self, lines: List[str]) -> str:
        """Determine the type of table structure"""
        if not lines:
            return 'unknown'

        # Check for tab separation
        tab_lines = sum(1 for line in lines if '\t' in line)
        if tab_lines > len(lines) * 0.5:
            return 'tab_separated'

        # Check for fixed-width columns
        if self._check_fixed_width_structure(lines):
            return 'fixed_width'

        # Check for space-separated
        return 'space_separated'

    def _check_column_consistency(self, lines: List[str]) -> float:
        """Check how consistent the column structure is"""
        if len(lines) < 2:
            return 0.0

        # Simple check: do most lines have similar number of fields?
        field_counts = [len(line.split()) for line in lines]
        if not field_counts:
            return 0.0

        most_common_count = max(set(field_counts), key=field_counts.count)
        consistency = field_counts.count(most_common_count) / len(field_counts)
        return consistency

    def _check_fixed_width_structure(self, lines: List[str]) -> bool:
        """Check if lines follow a fixed-width column structure"""
        if len(lines) < 3:
            return False

        # Look for consistent spacing patterns
        # This is a simplified check - could be enhanced
        return False  # Placeholder for now

class SemanticFieldMapper:
    """Maps detected fields to standard financial document schema"""

    def __init__(self, engine: PDFIntelligenceEngine):
        self.engine = engine
        self.standard_schema = {
            'Date': FieldType.DATE,
            'Description': FieldType.DESCRIPTION,
            'Reference': FieldType.REFERENCE,
            'Debit Amount': FieldType.DEBIT,
            'Credit Amount': FieldType.CREDIT,
            'Balance': FieldType.BALANCE,
            'Customer': FieldType.CUSTOMER,
            'Vendor': FieldType.VENDOR,
            'Invoice Number': FieldType.INVOICE_NUMBER,
            'Tax': FieldType.TAX
        }

    async def discover_fields(self, table_lines: List[str]) -> List[FieldMapping]:
        """Discover and map fields from table headers and data"""
        if not table_lines:
            return []

        # Find header line (usually first line or line with most keywords)
        header_line = self._find_header_line(table_lines)
        if not header_line:
            return []

        # Extract column headers
        headers = self._extract_headers(header_line)
        logger.info(f"🔍 Extracted headers: {headers}")

        # Get sample data for each column
        sample_data = self._extract_sample_data(table_lines, len(headers))

        # Map each header to a field type
        field_mappings = []
        for i, header in enumerate(headers):
            sample_values = sample_data[i] if i < len(sample_data) else []
            field_type = self.engine.detect_field_type(header, sample_values)
            standard_name = self._map_to_standard_name(field_type, header)

            mapping = FieldMapping(
                original_name=header,
                field_type=field_type,
                standard_name=standard_name,
                confidence=0.8,  # Base confidence
                sample_values=sample_values[:3]  # Keep first 3 samples
            )
            field_mappings.append(mapping)
            logger.info(f"📋 Mapped '{header}' → '{standard_name}' ({field_type.value})")

        return field_mappings

    def _find_header_line(self, lines: List[str]) -> Optional[str]:
        """Find the line that contains column headers"""
        header_keywords = ['date', 'description', 'amount', 'debit', 'credit', 'balance',
                          'reference', 'invoice', 'customer', 'vendor', 'tax']

        best_line = None
        best_score = 0

        for line in lines[:5]:  # Check first 5 lines
            line_lower = line.lower()
            score = sum(1 for keyword in header_keywords if keyword in line_lower)
            if score > best_score:
                best_score = score
                best_line = line

        return best_line if best_score >= 2 else lines[0]  # Fallback to first line

    def _extract_headers(self, header_line: str) -> List[str]:
        """Extract column headers from header line"""
        logger.info(f"🔍 Raw header line: '{header_line}'")

        # Try different separators
        if '\t' in header_line:
            headers = header_line.split('\t')
        elif '|' in header_line:
            headers = header_line.split('|')
        else:
            # Space-separated, but be smart about it
            headers = re.split(r'\s{2,}', header_line)  # Split on 2+ spaces
            if len(headers) == 1:
                headers = header_line.split()  # Fallback to single space

        headers = [h.strip() for h in headers if h.strip()]
        logger.info(f"🔍 Extracted headers: {headers}")
        return headers

    def _extract_sample_data(self, lines: List[str], num_columns: int) -> List[List[str]]:
        """Extract sample data for each column"""
        sample_data = [[] for _ in range(num_columns)]

        for line in lines[1:6]:  # Skip header, take next 5 lines
            if self._is_data_line(line):
                columns = self._split_line_into_columns(line, num_columns)
                for i, col_value in enumerate(columns):
                    if i < len(sample_data) and col_value.strip():
                        sample_data[i].append(col_value.strip())

        return sample_data

    def _is_data_line(self, line: str) -> bool:
        """Check if line contains data (not header or separator)"""
        line = line.strip()
        if not line or len(line) < 5:
            return False

        # Check for date pattern (strong indicator of data line)
        has_date = bool(re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', line))
        has_amount = bool(re.search(r'[\d,]+\.?\d*', line))

        return has_date or has_amount

    def _split_line_into_columns(self, line: str, expected_columns: int) -> List[str]:
        """Split line into columns intelligently"""
        logger.debug(f"🔍 Splitting line: '{line}' (expecting {expected_columns} columns)")

        if '\t' in line:
            columns = line.split('\t')
        elif '|' in line:
            columns = line.split('|')
        else:
            # For bank statements, use a more sophisticated approach
            # The PDF structure is: Date Description Reference [Debit|Credit] Balance
            # Some lines have 4 columns (Date Desc Ref Balance), others have 5 (Date Desc Ref Amount Balance)

            # First, try to split on multiple spaces to get natural columns
            natural_columns = re.split(r'\s{2,}', line)
            if len(natural_columns) < 2:
                # Fallback to single space split
                natural_columns = line.split()

            logger.debug(f"🔍 Natural columns: {natural_columns}")

            # Look for date pattern at start
            date_match = re.match(r'^(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line)
            if date_match and len(natural_columns) >= 4:
                date_part = natural_columns[0]

                # Find all numeric values (potential amounts)
                amounts = []
                amount_positions = []
                for i, col in enumerate(natural_columns):
                    if re.match(r'^[\d,]+\.?\d*$', col.replace(',', '')):
                        amounts.append(col)
                        amount_positions.append(i)

                logger.debug(f"🔍 Found amounts: {amounts} at positions: {amount_positions}")

                if len(amounts) == 1:
                    # Only one amount - this is either opening balance or single transaction
                    if len(natural_columns) == 4:
                        # Format: Date Description Reference Balance
                        columns = [natural_columns[0], natural_columns[1], natural_columns[2], "", "", natural_columns[3]]
                    else:
                        # Format: Date Description Reference Amount Balance - need to determine if Amount is debit or credit
                        amount = amounts[0]
                        balance = natural_columns[-1] if natural_columns[-1] != amount else ""

                        # Get description to determine transaction type
                        desc_parts = natural_columns[1:-2] if balance else natural_columns[1:-1]
                        description = ' '.join(desc_parts[:-1]) if len(desc_parts) > 1 else ' '.join(desc_parts)
                        reference = desc_parts[-1] if len(desc_parts) > 1 else ""

                        # Determine if it's debit or credit based on reference pattern first, then description
                        if reference.startswith('DEP'):
                            # DEP = Deposit = Credit
                            columns = [date_part, description, reference, "", amount, balance]
                        elif reference.startswith('CHQ') or reference.startswith('AUTO'):
                            # CHQ = Check, AUTO = Automatic charge = Debit
                            columns = [date_part, description, reference, amount, "", balance]
                        elif reference.startswith('OB'):
                            # Opening Balance - no debit/credit, just balance
                            columns = [date_part, description, reference, "", "", amount]
                        else:
                            # Fallback to description-based classification
                            debit_keywords = ['payment', 'expense', 'bill', 'charge', 'purchase', 'rent', 'utility', 'insurance', 'license', 'supplies', 'marketing', 'equipment', 'freelancer']
                            credit_keywords = ['received', 'deposit', 'revenue', 'income', 'retainer', 'consulting', 'service', 'client']

                            desc_lower = description.lower()
                            if any(keyword in desc_lower for keyword in debit_keywords):
                                columns = [date_part, description, reference, amount, "", balance]
                            elif any(keyword in desc_lower for keyword in credit_keywords):
                                columns = [date_part, description, reference, "", amount, balance]
                            else:
                                # Default to debit if uncertain
                                columns = [date_part, description, reference, amount, "", balance]

                elif len(amounts) == 2:
                    # Two amounts: likely Debit and Balance or Credit and Balance
                    # The last amount is usually the balance
                    balance = amounts[-1]
                    transaction_amount = amounts[0]

                    # Extract description and reference
                    desc_parts = []
                    for col in natural_columns[1:]:
                        if col not in amounts:
                            desc_parts.append(col)

                    if len(desc_parts) > 1:
                        description = ' '.join(desc_parts[:-1])
                        reference = desc_parts[-1]
                    else:
                        description = ' '.join(desc_parts)
                        reference = ""

                    # Determine if transaction amount is debit or credit based on reference pattern first
                    if reference.startswith('DEP'):
                        # DEP = Deposit = Credit
                        columns = [date_part, description, reference, "", transaction_amount, balance]
                    elif reference.startswith('CHQ') or reference.startswith('AUTO'):
                        # CHQ = Check, AUTO = Automatic charge = Debit
                        columns = [date_part, description, reference, transaction_amount, "", balance]
                    elif reference.startswith('OB'):
                        # Opening Balance - no debit/credit, just balance
                        columns = [date_part, description, reference, "", "", balance]
                    else:
                        # Fallback to description-based classification
                        debit_keywords = ['payment', 'expense', 'bill', 'charge', 'purchase', 'rent', 'utility', 'insurance', 'license', 'supplies', 'marketing', 'equipment', 'freelancer']
                        credit_keywords = ['received', 'deposit', 'revenue', 'income', 'retainer', 'consulting', 'service', 'client']

                        desc_lower = description.lower()
                        if any(keyword in desc_lower for keyword in debit_keywords):
                            columns = [date_part, description, reference, transaction_amount, "", balance]
                        elif any(keyword in desc_lower for keyword in credit_keywords):
                            columns = [date_part, description, reference, "", transaction_amount, balance]
                        else:
                            # Default to debit if uncertain
                            columns = [date_part, description, reference, transaction_amount, "", balance]

                else:
                    # No amounts or too many amounts, use natural split
                    columns = natural_columns
            else:
                # Not a transaction line, use natural split
                columns = natural_columns

        # Clean and pad columns
        columns = [col.strip() for col in columns]

        # Pad with empty strings if needed
        while len(columns) < expected_columns:
            columns.append('')

        return columns[:expected_columns]  # Truncate if too many

    def _map_to_standard_name(self, field_type: FieldType, original_name: str) -> str:
        """Map field type to standard schema name"""
        type_to_name = {
            FieldType.DATE: 'Date',
            FieldType.DESCRIPTION: 'Description',
            FieldType.REFERENCE: 'Reference',
            FieldType.DEBIT: 'Debit Amount',
            FieldType.CREDIT: 'Credit Amount',
            FieldType.BALANCE: 'Balance',
            FieldType.AMOUNT: 'Amount',
            FieldType.CUSTOMER: 'Customer',
            FieldType.VENDOR: 'Vendor',
            FieldType.INVOICE_NUMBER: 'Invoice Number',
            FieldType.TAX: 'Tax',
            FieldType.UNKNOWN: original_name  # Keep original if unknown
        }
        return type_to_name.get(field_type, original_name)

class DataExtractionOrchestrator:
    """Orchestrates data extraction using discovered field mappings"""

    def __init__(self, engine: PDFIntelligenceEngine, mapper: SemanticFieldMapper):
        self.engine = engine
        self.mapper = mapper

    async def extract_data(self, table_region: TableRegion, field_mappings: List[FieldMapping]) -> ExtractionResult:
        """Extract structured data using field mappings"""
        try:
            records = []
            headers = [mapping.standard_name for mapping in field_mappings]

            # Process each data line
            for line_idx, line in enumerate(table_region.lines[1:]):  # Skip header
                if not self.mapper._is_data_line(line):
                    continue

                # Split line into columns
                columns = self.mapper._split_line_into_columns(line, len(field_mappings))
                logger.debug(f"🔍 Line: '{line}' → Columns: {columns}")

                # Create record data
                record_data = {}
                for i, mapping in enumerate(field_mappings):
                    if i < len(columns):
                        raw_value = columns[i].strip()
                        processed_value = self._process_field_value(raw_value, mapping.field_type)
                        record_data[mapping.standard_name] = processed_value
                        logger.debug(f"🔍 Mapped '{raw_value}' → '{mapping.standard_name}': {processed_value}")

                # Add metadata
                record_data['id'] = f"pdf-record-{len(records)}"
                record_data['confidence'] = 0.85
                record_data['source'] = 'universal_pdf_extraction'

                # Create ExtractedRecord
                records.append(ExtractedRecord(
                    row_index=len(records),
                    data=record_data,
                    confidence=0.85
                ))

            logger.info(f"📊 Extracted {len(records)} records using universal processor")

            return ExtractionResult(
                records=records,
                total_records=len(records),
                extraction_confidence=0.85 if records else 0.3,
                schema_detected={
                    'format': 'universal_pdf',
                    'table_type': table_region.table_type,
                    'headers': ', '.join(headers)
                },
                processing_notes=[f"Universal extraction: {len(records)} records from {table_region.table_type} table"]
            )

        except Exception as e:
            logger.error(f"Data extraction failed: {e}")
            return ExtractionResult(
                records=[],
                total_records=0,
                extraction_confidence=0.0,
                schema_detected={},
                processing_notes=[f"Data extraction failed: {str(e)}"]
            )

    def _process_field_value(self, value: str, field_type: FieldType) -> Optional[str]:
        """Process field value based on its type"""
        if not value or value.lower() in ['', '-', 'nil', 'null']:
            return None

        if field_type == FieldType.AMOUNT or field_type == FieldType.DEBIT or field_type == FieldType.CREDIT or field_type == FieldType.BALANCE:
            # Clean amount values
            cleaned = re.sub(r'[^\d.,]', '', value)  # Remove currency symbols
            if cleaned:
                return cleaned

        return value

async def _fallback_line_processing(text: str, engine: PDFIntelligenceEngine) -> ExtractionResult:
    """Fallback processing when no table structure is detected"""
    try:
        lines = text.split('\n')
        records = []

        # Use basic pattern matching as fallback
        for line_idx, line in enumerate(lines):
            line = line.strip()
            if not line or len(line) < 10:
                continue

            # Look for lines with date and amount patterns
            date_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line)
            amount_matches = re.findall(r'([\d,]+\.?\d*)', line)

            if date_match and amount_matches:
                # Extract description (text between date and first amount)
                date_end = date_match.end()
                first_amount_start = line.find(amount_matches[0], date_end)
                description = line[date_end:first_amount_start].strip()

                # Create basic record
                record_data = {
                    'Date': date_match.group(1),
                    'Description': description,
                    'Amount': amount_matches[0],
                    'id': f"fallback-record-{len(records)}",
                    'confidence': 0.6,
                    'source': 'fallback_extraction'
                }

                records.append(ExtractedRecord(
                    row_index=len(records),
                    data=record_data,
                    confidence=0.6
                ))

        logger.info(f"📋 Fallback processing extracted {len(records)} records")

        return ExtractionResult(
            records=records,
            total_records=len(records),
            extraction_confidence=0.6 if records else 0.2,
            schema_detected={'format': 'fallback_extraction'},
            processing_notes=[f"Fallback extraction: {len(records)} records"]
        )

    except Exception as e:
        logger.error(f"Fallback processing failed: {e}")
        return ExtractionResult(
            records=[],
            total_records=0,
            extraction_confidence=0.0,
            schema_detected={},
            processing_notes=[f"Fallback processing failed: {str(e)}"]
        )

# Backward compatibility: alias the universal processor
async def process_bank_statement_pdf(content: bytes, filename: str) -> ExtractionResult:
    """Backward compatibility wrapper - uses universal PDF processor"""
    logger.info("🔄 Using universal PDF processor for backward compatibility")
    return await process_pdf_document(content, filename)

async def process_excel_register(content: bytes, filename: str, doc_type: str) -> ExtractionResult:
    """Process Excel sales/purchase register with structured data extraction"""
    try:
        excel_data = await extract_data_from_excel(content, filename)

        if excel_data.get('error'):
            raise Exception(excel_data['error'])

        headers = excel_data['headers']
        data_rows = excel_data['data']

        # Map common column variations
        column_mappings = {
            'sales_register': {
                'date': ['date', 'invoice_date', 'bill_date', 'transaction_date'],
                'customer': ['customer', 'customer_name', 'client', 'party_name'],
                'invoice_no': ['invoice_no', 'invoice_number', 'bill_no', 'voucher_no'],
                'amount': ['amount', 'total_amount', 'invoice_amount', 'total'],
                'tax': ['tax', 'gst', 'vat', 'tax_amount']
            },
            'purchase_register': {
                'date': ['date', 'purchase_date', 'bill_date', 'transaction_date'],
                'vendor': ['vendor', 'vendor_name', 'supplier', 'party_name'],
                'bill_no': ['bill_no', 'invoice_no', 'voucher_no', 'reference_no'],
                'amount': ['amount', 'total_amount', 'bill_amount', 'total'],
                'tax': ['tax', 'gst', 'vat', 'tax_amount']
            }
        }

        # Find column indices
        mappings = column_mappings.get(doc_type, {})
        column_indices = {}

        for field, variations in mappings.items():
            for i, header in enumerate(headers):
                if any(var.lower() in header.lower() for var in variations):
                    column_indices[field] = i
                    break

        records = []
        for row_idx, row in enumerate(data_rows):
            if len(row) > 0 and any(cell.strip() for cell in row if isinstance(cell, str)):
                record_data = {}

                # Map known columns
                for field, col_idx in column_indices.items():
                    if col_idx < len(row):
                        record_data[field] = row[col_idx]

                # Include all columns with headers
                for i, header in enumerate(headers):
                    if i < len(row):
                        record_data[header] = row[i]

                records.append(ExtractedRecord(
                    row_index=row_idx,
                    data=record_data,
                    confidence=0.9
                ))

        return ExtractionResult(
            records=records,
            total_records=len(records),
            extraction_confidence=0.9 if records else 0.3,
            schema_detected={'format': 'excel_register', 'columns': ', '.join(headers)},
            processing_notes=[f"Processed Excel {doc_type} with {len(records)} records"]
        )

    except Exception as e:
        logger.error(f"Excel register processing failed: {e}")
        return ExtractionResult(
            records=[],
            total_records=0,
            extraction_confidence=0.0,
            schema_detected={},
            processing_notes=[f"Excel processing failed: {str(e)}"]
        )

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "Finance AI Document Processor",
        "version": "1.0.0",
        "timestamp": datetime.now().isoformat(),
        "openai_available": openai_client is not None
    }

@app.post("/classify/document", response_model=ClassificationResult)
async def classify_document(file: UploadFile = File(...)):
    """Classify document type with confidence scoring"""
    try:
        # Extract text content
        content = await extract_text_from_file(file)
        
        # Classify using AI
        classification = await classify_document_with_ai(
            content=content,
            filename=file.filename or "unknown",
            file_type=file.content_type or "unknown"
        )
        
        return classification
        
    except Exception as e:
        logger.error(f"Error in document classification: {e}")
        raise HTTPException(status_code=500, detail=f"Classification failed: {str(e)}")

@app.post("/extract/document", response_model=DocumentProcessingResponse)
async def extract_document_data(file: UploadFile = File(...)):
    """Enhanced AI document processing with multi-format support"""
    start_time = datetime.now()
    logger.info(f"🔍 Starting document extraction for: {file.filename}")

    try:
        # Validate file format
        file_format = detect_file_format(file)
        logger.info(f"📄 Detected file format: {file_format}")
        if file_format == 'unknown':
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload PDF, Excel (.xlsx/.xls), or CSV files.")

        # Read file content
        content = await file.read()
        await file.seek(0)  # Reset file pointer for text extraction
        logger.info(f"📊 File size: {len(content)} bytes")

        # Extract text content for classification
        logger.info("🔤 Extracting text content for classification...")
        text_content = await extract_text_from_file(file)
        logger.info(f"📝 Extracted text length: {len(text_content)} characters")

        # Classify document
        logger.info("🤖 Starting AI document classification...")
        classification = await classify_document_with_ai(
            content=text_content,
            filename=file.filename or "unknown",
            file_type=file.content_type or "unknown"
        )
        logger.info(f"📋 Classification result: {classification.document_type} (confidence: {classification.confidence})")

        # Use format-specific processing based on file type and classification
        logger.info("⚙️ Starting data extraction...")
        if file_format == 'pdf' and classification.document_type == 'bank_statement':
            logger.info("🏦 Processing as PDF bank statement")
            extraction = await process_bank_statement_pdf(content, file.filename or "")
        elif file_format == 'excel' and classification.document_type in ['sales_register', 'purchase_register']:
            logger.info(f"📊 Processing as Excel {classification.document_type}")
            extraction = await process_excel_register(content, file.filename or "", classification.document_type)
        else:
            logger.info(f"🔧 Using general AI extraction for {classification.document_type}")
            extraction = await extract_data_with_ai(text_content, classification.document_type)

        logger.info(f"✅ Extraction completed: {len(extraction.records)} records extracted")

        # Calculate processing time
        processing_time = int((datetime.now() - start_time).total_seconds() * 1000)
        logger.info(f"⏱️ Total processing time: {processing_time}ms")

        return DocumentProcessingResponse(
            classification=classification,
            extraction=extraction,
            processing_time_ms=processing_time,
            ai_enhanced=openai_client is not None,
            raw_text_content=text_content
        )
        
    except Exception as e:
        logger.error(f"Error in document processing: {e}")
        raise HTTPException(status_code=500, detail=f"Document processing failed: {str(e)}")

# Authentication Endpoints
@app.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest, db = Depends(get_db)):
    """User authentication"""
    async with db.acquire() as conn:
        user_row = await conn.fetchrow(
            "SELECT id, email, first_name, last_name, password_hash, tenant_id FROM users WHERE email = $1",
            request.email
        )

        if not user_row or not verify_password(request.password, user_row['password_hash']):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid email or password"
            )

        token = generate_token(user_row['id'])

        # Update user token in database
        await conn.execute(
            "UPDATE users SET token = $1, updated_at = $2 WHERE id = $3",
            token, datetime.utcnow(), user_row['id']
        )

        user = User.from_db_row(user_row)

        return LoginResponse(token=token, user=user)

@app.get("/api/auth/user", response_model=User)
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    """Get current user information"""
    return current_user

# Dashboard Endpoints
@app.get("/api/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user), db = Depends(get_db)):
    """Get dashboard statistics"""
    async with db.acquire() as conn:
        # Get document count for user
        doc_count = await conn.fetchval(
            "SELECT COUNT(*) FROM documents WHERE uploaded_by = $1 AND tenant_id = $2",
            current_user.id, current_user.tenant_id
        )

        # Get recent processing errors
        error_count = await conn.fetchval(
            "SELECT COUNT(*) FROM documents WHERE uploaded_by = $1 AND tenant_id = $2 AND status = 'failed'",
            current_user.id, current_user.tenant_id
        )

        # Calculate compliance score (simplified)
        total_docs = doc_count or 1
        successful_docs = total_docs - (error_count or 0)
        compliance_score = int((successful_docs / total_docs) * 100)

        return {
            "documentsProcessed": doc_count or 0,
            "activeAgents": 7,  # Number of AI agents available
            "validationErrors": error_count or 0,
            "complianceScore": compliance_score
        }

# Workflow Endpoints
@app.get("/api/workflows")
async def get_workflows(current_user: User = Depends(get_current_user)):
    """Get available workflows"""
    # Return mock workflow data for now
    return [
        {
            "documentId": "sample-doc-1",
            "currentNode": "ClassifierBot",
            "nodes": {
                "ClassifierBot": {
                    "id": "ClassifierBot",
                    "name": "Document Classifier",
                    "status": "completed",
                    "type": "agent",
                    "dependencies": []
                },
                "ExtractorBot": {
                    "id": "ExtractorBot",
                    "name": "Data Extractor",
                    "status": "running",
                    "type": "agent",
                    "dependencies": ["ClassifierBot"]
                }
            },
            "completed": False
        }
    ]

# Audit Trail Endpoints
@app.get("/api/audit-trail")
async def get_audit_trail(current_user: User = Depends(get_current_user), db = Depends(get_db)):
    """Get audit trail for user's tenant"""
    async with db.acquire() as conn:
        # Get recent audit trail entries (mock data for now)
        return [
            {
                "id": "1",
                "action": "Document uploaded",
                "entityType": "document",
                "entityId": "doc-001",
                "userId": current_user.id,
                "details": {
                    "description": "Bank statement uploaded and processed successfully",
                    "fileName": "bank_statement_jan.pdf"
                },
                "timestamp": datetime.now().isoformat()
            },
            {
                "id": "2",
                "action": "ClassifierBot completed",
                "entityType": "agent_job",
                "entityId": "job-001",
                "userId": current_user.id,
                "details": {
                    "description": "Document classified as bank_statement with 95% confidence",
                    "confidence": 0.95
                },
                "timestamp": datetime.now().isoformat()
            }
        ]

# Compliance Endpoints
@app.get("/api/compliance-checks")
async def get_compliance_checks(current_user: User = Depends(get_current_user)):
    """Get compliance checks"""
    return [
        {
            "id": "1",
            "type": "GST_VALIDATION",
            "status": "passed",
            "description": "GST number validation completed",
            "timestamp": datetime.now().isoformat()
        },
        {
            "id": "2",
            "type": "BANK_RECONCILIATION",
            "status": "pending",
            "description": "Bank reconciliation in progress",
            "timestamp": datetime.now().isoformat()
        }
    ]

# Reports Endpoints
@app.post("/api/reports/trial-balance")
async def get_trial_balance(
    request: dict,
    current_user: User = Depends(get_current_user),
    db = Depends(get_db)
):
    """Get trial balance report"""
    period = request.get("period", "2025")

    # Mock trial balance data
    return {
        "period": period,
        "totalDebitsText": "Rs 4,75,689",
        "totalCreditsText": "Rs 4,75,689",
        "isBalanced": True,
        "accounts": [
            {"name": "Cash", "debit": 50000, "credit": 0},
            {"name": "Accounts Receivable", "debit": 75000, "credit": 0},
            {"name": "Accounts Payable", "debit": 0, "credit": 45000},
            {"name": "Revenue", "debit": 0, "credit": 80000}
        ],
        "generatedAt": datetime.now().isoformat()
    }

# Document Management Endpoints
@app.get("/api/documents")
async def get_documents(current_user: User = Depends(get_current_user), db = Depends(get_db)):
    """Get user's documents"""
    async with db.acquire() as conn:
        documents = await conn.fetch(
            """
            SELECT id, file_name, original_name, mime_type, file_size,
                   document_type, status, metadata, extracted_data,
                   created_at, updated_at
            FROM documents
            WHERE uploaded_by = $1 AND tenant_id = $2
            ORDER BY created_at DESC
            """,
            current_user.id, current_user.tenant_id
        )


        return [record_to_camel(doc) for doc in documents]


@app.get("/api/documents/{document_id}")
async def get_document(document_id: str, current_user: User = Depends(get_current_user), db = Depends(get_db)):
    """Get a single document by ID"""
    async with db.acquire() as conn:
        doc = await conn.fetchrow(
            """
            SELECT id, file_name, original_name, mime_type, file_size,
                   file_path, document_type, status, metadata, extracted_data,
                   created_at, updated_at
            FROM documents
            WHERE id = $1 AND uploaded_by = $2 AND tenant_id = $3
            """,
            document_id, current_user.id, current_user.tenant_id
        )
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        return record_to_camel(doc)


@app.delete("/api/documents/{document_id}")
async def delete_document(document_id: str, current_user: User = Depends(get_current_user), db = Depends(get_db)):
    """Delete a document and its file"""
    async with db.acquire() as conn:
        doc = await conn.fetchrow(
            "SELECT file_path FROM documents WHERE id = $1 AND uploaded_by = $2 AND tenant_id = $3",
            document_id, current_user.id, current_user.tenant_id
        )
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")

        await conn.execute(
            "DELETE FROM documents WHERE id = $1 AND uploaded_by = $2 AND tenant_id = $3",
            document_id, current_user.id, current_user.tenant_id
        )

    file_path = Path(doc["file_path"])
    if file_path.exists():
        try:
            file_path.unlink()
        except Exception as e:
            logger.warning(f"Failed to delete file {file_path}: {e}")

    return {"message": "Document deleted successfully"}


@app.get("/api/extracted-data")
async def get_extracted_data(period: Optional[str] = None, docType: Optional[str] = None,
                             current_user: User = Depends(get_current_user), db = Depends(get_db)):
    """Return extracted data records for use in data tables"""
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, document_type, file_name, metadata, extracted_data,
                   created_at, updated_at
            FROM documents
            WHERE uploaded_by = $1 AND tenant_id = $2
            ORDER BY created_at DESC
            """,
            current_user.id, current_user.tenant_id
        )

    docs = [dict(row) for row in rows]

    if period and period != "all":
        docs = [d for d in docs if (d.get("metadata") or {}).get("period", "Q1_2025") == period]

    extracted = []
    for d in docs:
        edata = d.get("extracted_data") or {}
        if not edata or "records" not in edata:
            continue
        extracted.append({
            "id": d["id"],
            "documentId": d["id"],
            "documentType": d["document_type"],
            "fileName": d["file_name"],
            "data": edata.get("records", []),
            "headers": edata.get("headers", []),
            "totalRecords": edata.get("total_records", 0),
            "extractedAt": edata.get("extractedAt") or d.get("updated_at") or d.get("created_at"),
            "confidence": 0.95,
            "extractionMethod": edata.get("extraction_method", "automated")
        })

    if docType and docType != "all":
        extracted = [e for e in extracted if e["documentType"] == docType]

    return extracted

@app.post("/api/documents/upload", response_model=DocumentUploadResponse)
async def upload_document(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db = Depends(get_db)
):
    """Upload and process document"""
    start_time = datetime.utcnow()

    # Validate file type by extension if MIME type is not reliable
    allowed_extensions = ['.pdf', '.csv', '.xlsx', '.xls']
    file_extension = Path(file.filename).suffix.lower()

    allowed_types = ['application/pdf', 'text/csv', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel', 'application/octet-stream']

    if file.content_type not in allowed_types and file_extension not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {file.content_type} (extension: {file_extension})")

    # Generate unique filename
    file_id = str(uuid.uuid4())
    file_extension = Path(file.filename).suffix
    unique_filename = f"{file_id}_{file.filename}"
    file_path = UPLOAD_DIR / unique_filename

    # Save file
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # Process document using existing endpoint logic
    # Reset file pointer for processing
    file.file.seek(0)

    # Use the existing document processing logic
    processing_result = await extract_document_data(file)
    classification = processing_result.classification
    extraction = processing_result.extraction

    # Create document record
    document_id = str(uuid.uuid4())
    document_data = {
        "id": document_id,
        "file_name": unique_filename,
        "original_name": file.filename,
        "mime_type": file.content_type,
        "file_size": file.size,
        "file_path": str(file_path),
        "document_type": classification.document_type,
        "status": "completed",
        "uploaded_by": current_user.id,
        "tenant_id": current_user.tenant_id,
        "metadata": {
            "classification": {
                "confidence": classification.confidence,
                "reasoning": classification.reasoning,
                "key_indicators": classification.key_indicators
            },
            "extraction": {
                "confidence": extraction.extraction_confidence,
                "total_records": extraction.total_records,
                "processing_notes": extraction.processing_notes
            }
        },
        "extracted_data": {
            "records": [record.model_dump() for record in extraction.records],
            "total_records": extraction.total_records,
            "confidence": extraction.extraction_confidence,
            "schema_detected": extraction.schema_detected
        },
        "created_at": start_time,
        "updated_at": start_time
    }

    # Save to database
    async with db.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO documents (
                id, file_name, original_name, mime_type, file_size, file_path,
                document_type, status, uploaded_by, tenant_id, metadata,
                extracted_data, created_at, updated_at
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14)
            """,
            document_id, unique_filename, file.filename, file.content_type, file.size,
            str(file_path), classification.document_type, "completed", current_user.id,
            current_user.tenant_id, json.dumps(document_data["metadata"]),
            json.dumps(document_data["extracted_data"]), start_time, start_time
        )

    processing_time = int((datetime.utcnow() - start_time).total_seconds() * 1000)

    return DocumentUploadResponse(
        document=Document(**document_data),
        message="Document uploaded and processed successfully",
        processing_time_ms=processing_time
    )

# Frontend serving
@app.get("/")
async def serve_frontend():
    """Serve React frontend"""
    return FileResponse("static/index.html")

@app.get("/{path:path}")
async def serve_frontend_routes(path: str):
    """Serve React frontend for all routes (SPA routing)"""
    # API routes should return 404 if not found
    if path.startswith("api/") or path.startswith("auth/") or path.startswith("extract/"):
        raise HTTPException(status_code=404, detail="API endpoint not found")

    # Check if it's a static asset
    if path.startswith("assets/"):
        static_file_path = f"static/{path}"
        if Path(static_file_path).exists():
            return FileResponse(static_file_path)
        raise HTTPException(status_code=404, detail="Static file not found")

    # For all other routes, serve the React app (SPA routing)
    return FileResponse("static/index.html")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")
